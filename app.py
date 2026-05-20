import streamlit as st
import re
import requests
import time
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

st.set_page_config(
    page_title="SIM Deal Auditor",
    page_icon="📱",
    layout="wide"
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Mono:wght@400;500&family=Syne:wght@700;800&display=swap');

html, body, [class*="css"] { font-family: 'DM Mono', monospace; }
h1, h2, h3 { font-family: 'Syne', sans-serif; }
.stApp { background: #0e0e0e; color: #e8e8e8; }
.block-container { padding: 2rem 2.5rem; max-width: 1400px; }

.hero {
    background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%);
    border: 1px solid #1e4d8c;
    border-radius: 12px;
    padding: 2.5rem;
    margin-bottom: 2rem;
    position: relative;
    overflow: hidden;
}
.hero h1 { font-size: 2.4rem; margin: 0 0 0.5rem; color: #fff; letter-spacing: -1px; }
.hero p { color: #8899bb; margin: 0; font-size: 0.85rem; }
</style>
""", unsafe_allow_html=True)

# ── Constants ─────────────────────────────────────────────────────────────────
NETWORKS = ["Three", "Vodafone", "SMARTY", "TalkMobile", "VOXI", "iD Mobile", "Lebara", "GiffGaff", "Spusu", "O2", "Sky"]

URLS = {
    "uSwitch": "https://www.uswitch.com/mobiles/sim-only-deals/",
    "MoneySuperMarket": "https://www.moneysupermarket.com/mobile-phones/sim-only-deals/",
    "CompareTheMarket": "https://www.comparethemarket.com/mobile-phones/sim-only/",
}

CONTRACT_LABELS = {1: "1M", 12: "12M", 24: "24M"}

# ── Colour logic ──────────────────────────────────────────────────────────────
def gb_color(gb_val, all_vals):
    valid = [v for v in all_vals if v is not None]
    if not valid or gb_val is None:
        return None
    mx, mn = max(valid), min(valid)
    if mx == mn:
        return "00B050"
    ratio = (gb_val - mn) / (mx - mn)
    return "00B050" if ratio >= 0.66 else ("FFC000" if ratio >= 0.33 else "FF0000")

# ── ScrapingBee fetch ─────────────────────────────────────────────────────────
def fetch_with_scrapingbee(url, api_key):
    """Fetch a URL via ScrapingBee with JS rendering and premium proxies."""
    try:
        resp = requests.get(
            "https://app.scrapingbee.com/api/v1/",
            params={
                "api_key": api_key,
                "url": url,
                "render_js": "true",
                "premium_proxy": "true",
                "country_code": "gb",
                "wait": "3000",
                "block_ads": "true",
            },
            timeout=60,
        )
        status = resp.status_code
        if status == 200:
            return resp.text, None
        else:
            # Return error detail for logging
            try:
                err = resp.json()
                msg = err.get("message", str(err))
            except:
                msg = resp.text[:200]
            return None, f"HTTP {status}: {msg}"
    except Exception as e:
        return None, f"Exception: {e}"

# ── HTML parser ───────────────────────────────────────────────────────────────
def parse_deals_from_html(html, source, contract_months):
    deals = []
    if not html:
        return deals

    clean = re.sub(r'<script[^>]*>.*?</script>', '', html, flags=re.DOTALL | re.IGNORECASE)
    clean = re.sub(r'<style[^>]*>.*?</style>', '', clean, flags=re.DOTALL | re.IGNORECASE)
    clean = re.sub(r'<[^>]+>', ' ', clean)
    clean = re.sub(r'\s+', ' ', clean)

    price_pattern = r'£(\d+(?:\.\d{2})?)'
    gb_pattern = r'(\d+)\s*GB|(\bUnlimited\b)'
    network_pattern = '|'.join(re.escape(n) for n in NETWORKS)

    price_matches = list(re.finditer(price_pattern, clean))
    seen = set()

    for pm in price_matches:
        price_val = float(pm.group(1))
        if price_val < 4 or price_val > 25:
            continue

        window = clean[max(0, pm.start() - 100): pm.start() + 400]

        gb_match = re.search(gb_pattern, window, re.IGNORECASE)
        if not gb_match:
            continue

        if gb_match.group(2):
            gb, gb_num = "Unlimited", 99999
        else:
            gb = int(gb_match.group(1))
            gb_num = gb

        net_match = re.search(network_pattern, window, re.IGNORECASE)
        network = net_match.group(0).strip() if net_match else "Unknown"

        key = (source, network, price_val, str(gb))
        if key in seen:
            continue
        seen.add(key)

        deals.append({
            "source": source,
            "network": network,
            "price": price_val,
            "gb": gb,
            "gb_num": gb_num,
            "contract": contract_months,
        })

    return deals

# ── Main scrape orchestrator ──────────────────────────────────────────────────
def run_scrape(api_key, contract_lengths, price_min, price_max):
    all_deals = []

    for months in contract_lengths:
        label = CONTRACT_LABELS[months]
        st.session_state.scrape_log.append(f"▶ Scraping {label} contracts...")

        for source, url in URLS.items():
            st.session_state.scrape_log.append(f"  → Fetching {source}...")
            html, err = fetch_with_scrapingbee(url, api_key)

            if html:
                # Save first 500 chars of clean text for debug
                clean_preview = re.sub(r"<[^>]+>", " ", html)
                clean_preview = re.sub(r"\s+", " ", clean_preview)[:300]
                st.session_state.scrape_log.append(f"  📄 Preview: {clean_preview}")
                deals = parse_deals_from_html(html, source, months)
                deals = [d for d in deals if price_min <= d["price"] <= price_max]
                all_deals.extend(deals)
                st.session_state.scrape_log.append(f"  ✓ {source}: {len(deals)} deals found")
            else:
                st.session_state.scrape_log.append(f"  ✗ {source}: {err}")

            time.sleep(1)  # polite delay between requests

    return all_deals

# ── Excel builder ─────────────────────────────────────────────────────────────
def build_excel(all_deals, contract_lengths):
    wb = Workbook()
    wb.remove(wb.active)

    SOURCES = ["uSwitch", "MoneySuperMarket", "CompareTheMarket"]
    PRICE_RANGE = range(5, 21)

    header_fills = {
        "uSwitch":           PatternFill("solid", fgColor="1F497D"),
        "MoneySuperMarket":  PatternFill("solid", fgColor="17375E"),
        "CompareTheMarket":  PatternFill("solid", fgColor="243F60"),
    }
    white_font = Font(color="FFFFFF", bold=True, size=9)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    grey_fill = PatternFill("solid", fgColor="F2F2F2")
    green_fill = PatternFill("solid", fgColor="00B050")

    def get_fill(hex): return PatternFill("solid", fgColor=hex)

    for months in contract_lengths:
        label = CONTRACT_LABELS[months]
        ws = wb.create_sheet(title=f"{label} Market")

        lookup = {}
        for d in all_deals:
            if d["contract"] != months:
                continue
            src, net, price = d["source"], d["network"], int(d["price"])
            lookup.setdefault(src, {}).setdefault(net, {})[price] = d["gb"]

        table_width = 1 + len(NETWORKS)
        gap = 1
        starts = [1, table_width + gap + 1, 2 * (table_width + gap) + 1]

        for t_idx, source in enumerate(SOURCES):
            col_start = starts[t_idx]

            ws.merge_cells(start_row=1, start_column=col_start,
                           end_row=1, end_column=col_start + table_width - 1)
            cell = ws.cell(row=1, column=col_start, value=f"{source} {label} Market")
            cell.fill = header_fills[source]
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

            ws.cell(row=2, column=col_start, value="Price").font = Font(bold=True, size=8)
            ws.cell(row=2, column=col_start).fill = grey_fill
            ws.column_dimensions[get_column_letter(col_start)].width = 7

            for n_idx, network in enumerate(NETWORKS):
                c = col_start + 1 + n_idx
                cell = ws.cell(row=2, column=c, value=network)
                cell.font = Font(bold=True, size=8)
                cell.alignment = Alignment(horizontal="center")
                cell.fill = grey_fill
                ws.column_dimensions[get_column_letter(c)].width = 10

            for r_idx, price in enumerate(PRICE_RANGE):
                row = 3 + r_idx
                ws.cell(row=row, column=col_start, value=f"£{price}").font = Font(bold=True, size=8)
                ws.cell(row=row, column=col_start).alignment = Alignment(horizontal="center")

                gb_vals = []
                for network in NETWORKS:
                    gb = lookup.get(source, {}).get(network, {}).get(price)
                    if gb is not None and gb != "Unlimited":
                        try: gb_vals.append(int(gb))
                        except: pass

                for n_idx, network in enumerate(NETWORKS):
                    c = col_start + 1 + n_idx
                    gb = lookup.get(source, {}).get(network, {}).get(price)
                    cell = ws.cell(row=row, column=c)
                    cell.alignment = Alignment(horizontal="center")
                    cell.font = Font(size=8)
                    cell.border = border

                    if gb is None:
                        cell.value = ""
                    elif gb == "Unlimited":
                        cell.value = "Unlimited"
                        cell.fill = green_fill
                        cell.font = Font(size=8, color="FFFFFF")
                    else:
                        cell.value = gb
                        color = gb_color(int(gb), gb_vals)
                        if color:
                            cell.fill = get_fill(color)
                            cell.font = Font(size=8, color="FFFFFF" if color in ("FF0000", "00B050") else "000000")

            ws.row_dimensions[1].height = 18
            ws.row_dimensions[2].height = 15

        ws.freeze_panes = "B3"

    ws_sum = wb.create_sheet(title="Summary", index=0)
    ws_sum["A1"] = "SIM Deal Audit"
    ws_sum["A1"].font = Font(bold=True, size=14)
    ws_sum["A2"] = f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}"
    ws_sum["A2"].font = Font(size=9, color="888888")
    ws_sum["A4"] = f"Total deals found: {len(all_deals)}"
    ws_sum["A5"] = f"Sources: uSwitch, MoneySuperMarket, CompareTheMarket"
    ws_sum["A6"] = f"Contract lengths: {', '.join(CONTRACT_LABELS[m] for m in contract_lengths)}"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ── UI ────────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="hero">
    <h1>📱 SIM Deal Auditor</h1>
    <p>Automated price comparison across uSwitch · MoneySuperMarket · CompareTheMarket</p>
</div>
""", unsafe_allow_html=True)

if "scrape_log" not in st.session_state:
    st.session_state.scrape_log = []
if "deals" not in st.session_state:
    st.session_state.deals = []

# ── API Key input ─────────────────────────────────────────────────────────────
st.markdown("#### 🔑 ScrapingBee API Key")
api_key = st.text_input(
    "Paste your ScrapingBee API key",
    type="password",
    help="Get a free key at scrapingbee.com — 1,000 free credits, no card needed"
)

if api_key:
    st.success("✓ API key entered")

st.divider()

# ── Controls ──────────────────────────────────────────────────────────────────
col1, col2, col3 = st.columns([2, 1, 1])

with col1:
    st.markdown("#### Contract Lengths")
    c1, c2, c3 = st.columns(3)
    inc_1m  = c1.checkbox("1 Month",   value=True)
    inc_12m = c2.checkbox("12 Months", value=True)
    inc_24m = c3.checkbox("24 Months", value=True)

with col2:
    st.markdown("#### Price Range")
    price_min = st.number_input("Min £", min_value=4, max_value=20, value=5)
    price_max = st.number_input("Max £", min_value=5, max_value=25, value=20)

with col3:
    st.markdown("#### Run Audit")
    st.write("")
    run_btn = st.button("🔍 Start Audit", use_container_width=True, type="primary", disabled=not api_key)
    if st.session_state.deals:
        contract_lengths = [m for m, inc in [(1, inc_1m), (12, inc_12m), (24, inc_24m)] if inc]
        excel_data = build_excel(st.session_state.deals, contract_lengths)
        st.download_button(
            "⬇ Download Excel",
            data=excel_data,
            file_name=f"sim_audit_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

st.divider()

# ── Run ───────────────────────────────────────────────────────────────────────
if run_btn:
    contract_lengths = [m for m, inc in [(1, inc_1m), (12, inc_12m), (24, inc_24m)] if inc]
    if not contract_lengths:
        st.warning("Please select at least one contract length.")
    else:
        st.session_state.scrape_log = []
        st.session_state.deals = []

        # Credit estimate
        total_requests = len(contract_lengths) * 3
        st.info(f"ℹ️ This audit will use approximately **{total_requests * 75} credits** ({total_requests} pages × 75 credits each with JS + premium proxy).")

        with st.spinner("Scraping in progress — this takes 1-2 minutes..."):
            deals = run_scrape(api_key, contract_lengths, price_min, price_max)

        st.session_state.deals = deals
        st.success(f"✅ Audit complete — {len(deals)} deals collected")

# ── Log ───────────────────────────────────────────────────────────────────────
if st.session_state.scrape_log:
    with st.expander("📋 Scrape Log", expanded=True):
        for line in st.session_state.scrape_log:
            st.text(line)

# ── Results ───────────────────────────────────────────────────────────────────
if st.session_state.deals:
    df = pd.DataFrame(st.session_state.deals)
    st.markdown("### Results")

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Total Deals", len(df))
    m2.metric("uSwitch", len(df[df.source == "uSwitch"]))
    m3.metric("MoneySuperMarket", len(df[df.source == "MoneySuperMarket"]))
    m4.metric("CompareTheMarket", len(df[df.source == "CompareTheMarket"]))

    display_df = df[["source", "network", "price", "gb", "contract"]].copy()
    display_df.columns = ["Source", "Network", "Price (£)", "Data", "Contract (M)"]
    display_df = display_df.sort_values(["Source", "Price (£)", "Network"])
    st.dataframe(display_df, use_container_width=True, height=400)

    contract_lengths = [m for m, inc in [(1, inc_1m), (12, inc_12m), (24, inc_24m)] if inc]
    excel_data = build_excel(st.session_state.deals, contract_lengths)
    st.download_button(
        "⬇ Download Full Excel Report",
        data=excel_data,
        file_name=f"sim_audit_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
else:
    if not api_key:
        st.markdown("""
        <div style="text-align:center; padding:3rem; color:#6b7280;">
            <div style="font-size:2.5rem">🔑</div>
            <div style="margin-top:1rem">Enter your ScrapingBee API key above to get started</div>
            <div style="font-size:0.8rem; margin-top:0.5rem">Free at scrapingbee.com — no credit card needed</div>
        </div>
        """, unsafe_allow_html=True)
    else:
        st.markdown("""
        <div style="text-align:center; padding:3rem; color:#6b7280;">
            <div style="font-size:2.5rem">📡</div>
            <div style="margin-top:1rem">Select options above and click Start Audit</div>
        </div>
        """, unsafe_allow_html=True)
